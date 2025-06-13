#!/usr/bin/env python3
import openpyxl
import sys

def check_excel_format(excel_path):
    """Diagnostic tool to check Excel column contents."""
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        
        print(f"Sheet name: {sheet.title}")
        print(f"Max row: {sheet.max_row}")
        print(f"Max column: {sheet.max_column}")
        print("\nFirst 20 rows, columns A-Z:")
        print("-" * 100)
        
        for row in range(1, 21):
            row_data = []
            for col in range(1, 27):  # A-Z
                value = sheet.cell(row=row, column=col).value
                if value is not None:
                    # Truncate long values
                    str_val = str(value)[:20]
                    if len(str(value)) > 20:
                        str_val += "..."
                    row_data.append(f"{col}:{str_val}")
            if row_data:
                print(f"Row {row}: {' | '.join(row_data)}")
        
        print("\n" + "-" * 100)
        print("Checking row 15 specifically (where data should start):")
        for col in range(1, 30):
            value = sheet.cell(row=15, column=col).value
            if value is not None:
                print(f"  Column {col}: {value}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        check_excel_format(sys.argv[1])
    else:
        print("Usage: python check_excel_columns.py <excel_file_path>")