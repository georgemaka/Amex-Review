from typing import List, Any, Optional, Dict
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_, func, and_
from sqlalchemy.orm import selectinload
import openpyxl
import io

from app.core.security import get_current_user, check_user_role
from app.db.models import (
    User, UserRole, Cardholder, CardholderAssignment, 
    CardholderReviewer, CardholderStatement, Transaction, Statement
)
from app.db.schemas import (
    Cardholder as CardholderSchema,
    CardholderCreate,
    CardholderUpdate,
    CardholderAssignment as AssignmentSchema,
    CardholderAssignmentCreate,
    CardholderReviewer as ReviewerSchema,
    CardholderReviewerCreate
)
from app.db.session import get_async_db

router = APIRouter()


@router.get("/", response_model=List[CardholderSchema])
async def list_cardholders(
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(get_current_user),
    is_active: Optional[bool] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
) -> Any:
    query = select(Cardholder)
    
    # Filter by active status
    if is_active is not None:
        query = query.where(Cardholder.is_active == is_active)
    
    # Search by name
    if search:
        search_term = f"%{search}%"
        query = query.where(
            or_(
                Cardholder.full_name.ilike(search_term),
                Cardholder.first_name.ilike(search_term),
                Cardholder.last_name.ilike(search_term),
                Cardholder.employee_id.ilike(search_term)
            )
        )
    
    query = query.order_by(Cardholder.full_name).offset(skip).limit(limit)
    result = await db.execute(query)
    cardholders = result.scalars().all()
    
    return cardholders


@router.get("/with-transaction-counts", response_model=List[Dict[str, Any]])
async def list_cardholders_with_counts(
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(get_current_user),
    is_active: Optional[bool] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
) -> Any:
    """Get cardholders with their transaction counts."""
    # Base query for cardholders
    cardholder_query = select(Cardholder)
    
    # Filter by active status
    if is_active is not None:
        cardholder_query = cardholder_query.where(Cardholder.is_active == is_active)
    
    # Search by name
    if search:
        search_term = f"%{search}%"
        cardholder_query = cardholder_query.where(
            or_(
                Cardholder.full_name.ilike(search_term),
                Cardholder.first_name.ilike(search_term),
                Cardholder.last_name.ilike(search_term),
                Cardholder.employee_id.ilike(search_term)
            )
        )
    
    cardholder_query = cardholder_query.order_by(Cardholder.full_name).offset(skip).limit(limit)
    cardholder_result = await db.execute(cardholder_query)
    cardholders = cardholder_result.scalars().all()
    
    # Get transaction counts for each cardholder
    result = []
    for cardholder in cardholders:
        # Count transactions through cardholder_statements
        count_query = select(func.count(Transaction.id)).select_from(Transaction).join(
            CardholderStatement,
            Transaction.cardholder_statement_id == CardholderStatement.id
        ).where(CardholderStatement.cardholder_id == cardholder.id)
        
        count_result = await db.execute(count_query)
        transaction_count = count_result.scalar() or 0
        
        # Convert to dict and add count
        cardholder_dict = {
            "id": cardholder.id,
            "full_name": cardholder.full_name,
            "first_name": cardholder.first_name,
            "last_name": cardholder.last_name,
            "employee_id": cardholder.employee_id,
            "card_number": cardholder.card_number,
            "department": cardholder.department,
            "is_active": cardholder.is_active,
            "transaction_count": transaction_count
        }
        result.append(cardholder_dict)
    
    return result


@router.get("/with-assignments")
async def get_cardholders_with_assignments(
    is_active: bool = True,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(check_user_role([UserRole.ADMIN]))
) -> Any:
    """Get all cardholders with their current assignments (supports multiple assignments per cardholder)."""
    # Get all active cardholders
    cardholder_result = await db.execute(
        select(Cardholder)
        .where(Cardholder.is_active == is_active)
        .order_by(Cardholder.full_name)
    )
    cardholders = cardholder_result.scalars().all()
    
    # Get all coder assignments
    coder_assignments = await db.execute(
        select(
            CardholderAssignment.cardholder_id,
            CardholderAssignment.id,
            User.id,
            func.concat(User.first_name, ' ', User.last_name).label('coder_name')
        )
        .join(User, CardholderAssignment.coder_id == User.id)
        .where(CardholderAssignment.is_active == True)
    )
    
    # Get all reviewer assignments
    reviewer_assignments = await db.execute(
        select(
            CardholderReviewer.cardholder_id,
            CardholderReviewer.id,
            User.id,
            func.concat(User.first_name, ' ', User.last_name).label('reviewer_name')
        )
        .join(User, CardholderReviewer.reviewer_id == User.id)
        .where(CardholderReviewer.is_active == True)
    )
    
    # Build assignment mappings
    coder_map = {}
    for row in coder_assignments:
        cardholder_id = row[0]
        if cardholder_id not in coder_map:
            coder_map[cardholder_id] = []
        coder_map[cardholder_id].append({
            "id": row[2],
            "name": row[3],
            "assignment_id": row[1]
        })
    
    reviewer_map = {}
    for row in reviewer_assignments:
        cardholder_id = row[0]
        if cardholder_id not in reviewer_map:
            reviewer_map[cardholder_id] = []
        reviewer_map[cardholder_id].append({
            "id": row[2],
            "name": row[3],
            "assignment_id": row[1]
        })
    
    # Build response
    result = []
    for cardholder in cardholders:
        result.append({
            "id": cardholder.id,
            "full_name": cardholder.full_name,
            "first_name": cardholder.first_name,
            "last_name": cardholder.last_name,
            "department": cardholder.department,
            "is_active": cardholder.is_active,
            "assigned_coders": coder_map.get(cardholder.id, []),
            "assigned_reviewers": reviewer_map.get(cardholder.id, [])
        })
    
    return result


@router.post("/", response_model=CardholderSchema)
async def create_cardholder(
    cardholder_in: CardholderCreate,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(check_user_role([UserRole.ADMIN]))
) -> Any:
    # Check if cardholder already exists
    existing = await db.execute(
        select(Cardholder).where(Cardholder.full_name == cardholder_in.full_name)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(400, "Cardholder with this name already exists")
    
    cardholder = Cardholder(**cardholder_in.model_dump())
    db.add(cardholder)
    await db.commit()
    await db.refresh(cardholder)
    
    # Check if there are any cardholder_statements with matching name and no cardholder_id
    # Update them to link to the newly created cardholder
    cardholder_statements_result = await db.execute(
        select(CardholderStatement).where(
            CardholderStatement.cardholder_id == None
        )
    )
    cardholder_statements = cardholder_statements_result.scalars().all()
    
    updated_count = 0
    for cs in cardholder_statements:
        # Extract cardholder name from PDF path
        pdf_filename = cs.pdf_path.split('/')[-1]
        name_part = pdf_filename.replace('.pdf', '')
        parts = name_part.split('_', 2)
        if len(parts) >= 3:
            extracted_name = parts[2]
            # Check if the extracted name matches the newly created cardholder
            if extracted_name == cardholder.full_name:
                cs.cardholder_id = cardholder.id
                updated_count += 1
    
    if updated_count > 0:
        await db.commit()
    
    return cardholder


@router.get("/{cardholder_id}", response_model=CardholderSchema)
async def get_cardholder(
    cardholder_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    result = await db.execute(
        select(Cardholder).where(Cardholder.id == cardholder_id)
    )
    cardholder = result.scalar_one_or_none()
    
    if not cardholder:
        raise HTTPException(404, "Cardholder not found")
    
    return cardholder


@router.put("/{cardholder_id}", response_model=CardholderSchema)
async def update_cardholder(
    cardholder_id: int,
    cardholder_in: CardholderUpdate,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(check_user_role([UserRole.ADMIN]))
) -> Any:
    result = await db.execute(
        select(Cardholder).where(Cardholder.id == cardholder_id)
    )
    cardholder = result.scalar_one_or_none()
    
    if not cardholder:
        raise HTTPException(404, "Cardholder not found")
    
    # Update fields
    update_data = cardholder_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(cardholder, field, value)
    
    await db.commit()
    await db.refresh(cardholder)
    
    return cardholder


@router.delete("/{cardholder_id}")
async def delete_cardholder(
    cardholder_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(check_user_role([UserRole.ADMIN]))
) -> Any:
    result = await db.execute(
        select(Cardholder).where(Cardholder.id == cardholder_id)
    )
    cardholder = result.scalar_one_or_none()
    
    if not cardholder:
        raise HTTPException(404, "Cardholder not found")
    
    # Soft delete - just mark as inactive
    cardholder.is_active = False
    await db.commit()
    
    return {"message": "Cardholder deactivated"}


# Assignment endpoints
@router.get("/{cardholder_id}/assignments", response_model=List[AssignmentSchema])
async def get_cardholder_assignments(
    cardholder_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    result = await db.execute(
        select(CardholderAssignment)
        .options(selectinload(CardholderAssignment.coder))
        .where(CardholderAssignment.cardholder_id == cardholder_id)
    )
    assignments = result.scalars().all()
    
    return assignments


@router.post("/{cardholder_id}/assignments", response_model=AssignmentSchema)
async def create_cardholder_assignment(
    cardholder_id: int,
    assignment_in: CardholderAssignmentCreate,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(check_user_role([UserRole.ADMIN]))
) -> Any:
    # Verify cardholder exists
    cardholder_result = await db.execute(
        select(Cardholder).where(Cardholder.id == cardholder_id)
    )
    if not cardholder_result.scalar_one_or_none():
        raise HTTPException(404, "Cardholder not found")
    
    # Check if assignment already exists
    existing = await db.execute(
        select(CardholderAssignment).where(
            CardholderAssignment.cardholder_id == cardholder_id,
            CardholderAssignment.coder_id == assignment_in.coder_id,
            CardholderAssignment.is_active == True
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(400, "Active assignment already exists")
    
    assignment = CardholderAssignment(
        cardholder_id=cardholder_id,
        **assignment_in.model_dump(exclude={"cardholder_id"})
    )
    db.add(assignment)
    await db.commit()
    await db.refresh(assignment)
    
    # Load the relationships for the response
    result = await db.execute(
        select(CardholderAssignment)
        .options(
            selectinload(CardholderAssignment.cardholder),
            selectinload(CardholderAssignment.coder)
        )
        .where(CardholderAssignment.id == assignment.id)
    )
    assignment_with_relations = result.scalar_one()
    
    return assignment_with_relations


# Reviewer endpoints
@router.get("/{cardholder_id}/reviewers", response_model=List[ReviewerSchema])
async def get_cardholder_reviewers(
    cardholder_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    result = await db.execute(
        select(CardholderReviewer)
        .options(selectinload(CardholderReviewer.reviewer))
        .where(CardholderReviewer.cardholder_id == cardholder_id)
        .order_by(CardholderReviewer.review_order)
    )
    reviewers = result.scalars().all()
    
    return reviewers


@router.post("/{cardholder_id}/reviewers", response_model=ReviewerSchema)
async def create_cardholder_reviewer(
    cardholder_id: int,
    reviewer_in: CardholderReviewerCreate,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(check_user_role([UserRole.ADMIN]))
) -> Any:
    # Verify cardholder exists
    cardholder_result = await db.execute(
        select(Cardholder).where(Cardholder.id == cardholder_id)
    )
    if not cardholder_result.scalar_one_or_none():
        raise HTTPException(404, "Cardholder not found")
    
    # Check if reviewer already exists
    existing = await db.execute(
        select(CardholderReviewer).where(
            CardholderReviewer.cardholder_id == cardholder_id,
            CardholderReviewer.reviewer_id == reviewer_in.reviewer_id,
            CardholderReviewer.is_active == True
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(400, "Active reviewer assignment already exists")
    
    reviewer = CardholderReviewer(
        cardholder_id=cardholder_id,
        **reviewer_in.model_dump(exclude={"cardholder_id"})
    )
    db.add(reviewer)
    await db.commit()
    await db.refresh(reviewer)
    
    # Load the relationships for the response
    result = await db.execute(
        select(CardholderReviewer)
        .options(
            selectinload(CardholderReviewer.cardholder),
            selectinload(CardholderReviewer.reviewer)
        )
        .where(CardholderReviewer.id == reviewer.id)
    )
    reviewer_with_relations = result.scalar_one()
    
    return reviewer_with_relations


@router.delete("/assignments/{assignment_id}")
async def delete_cardholder_assignment(
    assignment_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(check_user_role([UserRole.ADMIN]))
) -> Any:
    """Delete (deactivate) a cardholder assignment."""
    result = await db.execute(
        select(CardholderAssignment).where(CardholderAssignment.id == assignment_id)
    )
    assignment = result.scalar_one_or_none()
    
    if not assignment:
        raise HTTPException(404, "Assignment not found")
    
    assignment.is_active = False
    await db.commit()
    
    return {"message": "Assignment removed successfully"}


@router.delete("/reviewers/{reviewer_id}")
async def delete_cardholder_reviewer(
    reviewer_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(check_user_role([UserRole.ADMIN]))
) -> Any:
    """Delete (deactivate) a cardholder reviewer assignment."""
    result = await db.execute(
        select(CardholderReviewer).where(CardholderReviewer.id == reviewer_id)
    )
    reviewer = result.scalar_one_or_none()
    
    if not reviewer:
        raise HTTPException(404, "Reviewer assignment not found")
    
    reviewer.is_active = False
    await db.commit()
    
    return {"message": "Reviewer assignment removed successfully"}


@router.get("/missing", response_model=List[Dict[str, Any]])
async def get_missing_cardholders(
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(check_user_role([UserRole.ADMIN]))
) -> Any:
    """Get cardholders that appear in statements but are not in the cardholder management system."""
    # Get all unique cardholder names from cardholder_statements that don't have a cardholder_id
    missing_query = select(
        CardholderStatement.id,
        CardholderStatement.statement_id,
        CardholderStatement.pdf_path,
        CardholderStatement.total_amount,
        CardholderStatement.transaction_count,
        Statement.month,
        Statement.year
    ).join(
        Statement, CardholderStatement.statement_id == Statement.id
    ).where(
        CardholderStatement.cardholder_id == None
    ).order_by(
        Statement.year.desc(),
        Statement.month.desc()
    )
    
    result = await db.execute(missing_query)
    missing_cardholder_statements = result.all()
    
    # Group by cardholder name extracted from PDF path
    missing_cardholders = {}
    
    for cs in missing_cardholder_statements:
        # Extract cardholder name from PDF path
        # Format: /app/data/statements/2024_01_Smith John.pdf
        pdf_filename = cs.pdf_path.split('/')[-1]
        # Remove extension and date prefix
        name_part = pdf_filename.replace('.pdf', '')
        # Extract name after year_month_
        parts = name_part.split('_', 2)
        if len(parts) >= 3:
            cardholder_name = parts[2]
            
            if cardholder_name not in missing_cardholders:
                # Parse name to get first and last name
                name_parts = cardholder_name.split(' ', 1)
                last_name = name_parts[0] if name_parts else cardholder_name
                first_name = name_parts[1] if len(name_parts) > 1 else ""
                
                missing_cardholders[cardholder_name] = {
                    "full_name": cardholder_name,
                    "first_name": first_name,
                    "last_name": last_name,
                    "statement_count": 0,
                    "total_transactions": 0,
                    "total_amount": 0,
                    "statements": []
                }
            
            missing_cardholders[cardholder_name]["statement_count"] += 1
            missing_cardholders[cardholder_name]["total_transactions"] += cs.transaction_count
            missing_cardholders[cardholder_name]["total_amount"] += cs.total_amount
            missing_cardholders[cardholder_name]["statements"].append({
                "id": cs.id,
                "statement_id": cs.statement_id,
                "month": cs.month,
                "year": cs.year,
                "transaction_count": cs.transaction_count,
                "amount": cs.total_amount
            })
    
    # Convert to list
    result = list(missing_cardholders.values())
    
    # Also check for cardholders in statements where the cardholder_id doesn't exist in cardholders table
    orphaned_query = select(
        CardholderStatement.cardholder_id,
        func.count(CardholderStatement.id).label('statement_count'),
        func.sum(CardholderStatement.transaction_count).label('total_transactions'),
        func.sum(CardholderStatement.total_amount).label('total_amount')
    ).where(
        and_(
            CardholderStatement.cardholder_id != None,
            ~CardholderStatement.cardholder_id.in_(
                select(Cardholder.id)
            )
        )
    ).group_by(CardholderStatement.cardholder_id)
    
    orphaned_result = await db.execute(orphaned_query)
    orphaned_cardholders = orphaned_result.all()
    
    # Add orphaned cardholders to result
    for orphan in orphaned_cardholders:
        result.append({
            "full_name": f"Unknown (ID: {orphan.cardholder_id})",
            "first_name": "",
            "last_name": f"Unknown (ID: {orphan.cardholder_id})",
            "cardholder_id": orphan.cardholder_id,
            "statement_count": orphan.statement_count,
            "total_transactions": orphan.total_transactions,
            "total_amount": orphan.total_amount,
            "statements": [],
            "is_orphaned": True
        })
    
    return result


@router.post("/import")
async def import_cardholders_from_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(check_user_role([UserRole.ADMIN]))
) -> Any:
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "File must be an Excel file")
    
    try:
        # Read Excel file
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = wb.active
        
        imported_count = 0
        errors = []
        
        # Process rows (assuming headers in row 1)
        for row_num in range(2, sheet.max_row + 1):
            try:
                # Extract data based on expected columns
                pdf_name = sheet.cell(row=row_num, column=1).value
                csv_name = sheet.cell(row=row_num, column=2).value
                coder_email = sheet.cell(row=row_num, column=3).value
                cc_email = sheet.cell(row=row_num, column=4).value
                
                if not pdf_name:
                    continue
                
                # Parse name
                name_parts = pdf_name.strip().split()
                first_name = name_parts[0] if name_parts else ""
                last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
                full_name = pdf_name.strip()
                
                # Get or create cardholder
                cardholder_result = await db.execute(
                    select(Cardholder).where(Cardholder.full_name == full_name)
                )
                cardholder = cardholder_result.scalar_one_or_none()
                
                if not cardholder:
                    cardholder = Cardholder(
                        full_name=full_name,
                        first_name=first_name,
                        last_name=last_name
                    )
                    db.add(cardholder)
                    await db.flush()
                
                # Create assignment if coder email provided
                if coder_email:
                    # Get coder user
                    coder_result = await db.execute(
                        select(User).where(User.email == coder_email)
                    )
                    coder = coder_result.scalar_one_or_none()
                    
                    if coder:
                        # Check if assignment exists
                        assignment_result = await db.execute(
                            select(CardholderAssignment).where(
                                CardholderAssignment.cardholder_id == cardholder.id,
                                CardholderAssignment.coder_id == coder.id
                            )
                        )
                        assignment = assignment_result.scalar_one_or_none()
                        
                        if not assignment:
                            cc_emails = [cc_email] if cc_email else []
                            assignment = CardholderAssignment(
                                cardholder_id=cardholder.id,
                                coder_id=coder.id,
                                cc_emails=cc_emails
                            )
                            db.add(assignment)
                    else:
                        errors.append(f"Row {row_num}: Coder {coder_email} not found")
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        await db.commit()
        
        return {
            "imported": imported_count,
            "errors": errors
        }
        
    except Exception as e:
        raise HTTPException(500, f"Failed to process file: {str(e)}")