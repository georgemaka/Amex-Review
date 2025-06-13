import os
import csv
from typing import Dict, List, Optional
from datetime import datetime
import openpyxl
import pandas as pd
import logging

from app.core.config import settings

logger = logging.getLogger(__name__)


class ExcelProcessor:
    def __init__(self):
        self.vendor_code = settings.AMEX_VENDOR_CODE
        self.jcco = "1"  # Job Cost Company
        self.record_type = "3"  # Type for APLB records
        self.header_row = None
        self.column_map = {}
    
    def _find_header_row(self, sheet) -> int:
        """Find the header row by looking for 'Product' in the first column."""
        for row_num in range(1, 30):  # Check first 30 rows
            cell_value = sheet.cell(row=row_num, column=1).value
            if cell_value and "Product" in str(cell_value):
                return row_num
        raise ValueError("Could not find header row with 'Product' column")
    
    def _map_columns(self, sheet, header_row: int) -> None:
        """Map column names to column numbers based on header row."""
        self.column_map = {}
        
        # Read all headers from the header row
        for col_num in range(1, sheet.max_column + 1):
            header = sheet.cell(row=header_row, column=col_num).value
            if header:
                # Clean header: remove newlines and extra spaces
                header_str = ' '.join(str(header).split()).strip()
                
                # Map the important columns
                if "Basic Card Account No" in header_str:
                    self.column_map['basic_card_account'] = col_num
                elif "Supplemental Cardmember Last Name" in header_str:
                    self.column_map['supp_last_name'] = col_num
                elif "Supplemental Cardmember First Name" in header_str:
                    self.column_map['supp_first_name'] = col_num
                elif "Supplemental Account Number" in header_str:
                    self.column_map['supp_card_number'] = col_num
                elif "Business Process Date" in header_str:
                    self.column_map['business_process_date'] = col_num
                elif "Transaction Date" in header_str and "Reference" not in header_str:
                    self.column_map['transaction_date'] = col_num
                elif "Transaction Amount USD" in header_str:
                    self.column_map['amount'] = col_num
                elif "Transaction Reference" in header_str:
                    self.column_map['transaction_reference'] = col_num
                    
                # Map description columns (up to 16)
                for i in range(1, 17):
                    if f"Transaction Description {i}" in header_str:
                        self.column_map[f'description_{i}'] = col_num
                        
        logger.info(f"Mapped columns: {self.column_map}")
        
        # Log all found headers for debugging
        all_headers = []
        for col_num in range(1, min(sheet.max_column + 1, 50)):
            header = sheet.cell(row=header_row, column=col_num).value
            if header:
                all_headers.append(f"Col {col_num}: {str(header)[:50]}")
        logger.info(f"All headers found: {all_headers}")
        
        # Validate required columns
        required_columns = ['amount', 'business_process_date', 'transaction_date']
        missing = [col for col in required_columns if col not in self.column_map]
        if missing:
            logger.error(f"Missing required columns: {missing}")
            logger.error(f"Available columns: {list(self.column_map.keys())}")
            raise ValueError(f"Missing required columns: {missing}. Please check Excel format.")
    
    def parse_statement(self, excel_path: str) -> Dict[str, List[Dict]]:
        """
        Parse AMEX Excel statement and group transactions by cardholder.
        Returns dict with cardholder names as keys and transaction lists as values.
        """
        transactions_by_cardholder = {}
        
        try:
            # Load Excel file
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = wb.active
            
            # Find header row and map columns
            self.header_row = self._find_header_row(sheet)
            self._map_columns(sheet, self.header_row)
            logger.info(f"Found header row at: {self.header_row}")
            
            # Start processing from the row after headers
            data_start_row = self.header_row + 1
            
            # Process data rows
            for row_num in range(data_start_row, sheet.max_row + 1):
                # Check if this row has transaction data by looking for amount
                amount_value = None
                if 'amount' in self.column_map:
                    amount_value = sheet.cell(row=row_num, column=self.column_map['amount']).value
                
                if amount_value is None:
                    continue
                
                # Extract cardholder information
                first_name = ""
                last_name = ""
                card_number = ""
                
                # Get supplemental cardholder info (primary for our use)
                if 'supp_first_name' in self.column_map:
                    first_name = self._clean_value(sheet.cell(row=row_num, column=self.column_map['supp_first_name']).value)
                if 'supp_last_name' in self.column_map:
                    last_name = self._clean_value(sheet.cell(row=row_num, column=self.column_map['supp_last_name']).value)
                if 'supp_card_number' in self.column_map:
                    card_number = self._clean_value(sheet.cell(row=row_num, column=self.column_map['supp_card_number']).value)
                
                # Skip if no cardholder name
                if not first_name or not last_name:
                    continue
                
                # Make names uppercase for consistency
                first_name = first_name.upper()
                last_name = last_name.upper()
                cardholder_name = f"{first_name} {last_name}"
                
                # Extract description/merchant from multiple columns
                desc_parts = []
                for i in range(1, 17):  # Check up to 16 description columns
                    desc_col = f'description_{i}'
                    if desc_col in self.column_map:
                        desc_val = self._clean_value(sheet.cell(row=row_num, column=self.column_map[desc_col]).value)
                        if desc_val:
                            desc_parts.append(desc_val)
                
                # Primary description is usually the merchant name
                merchant_name = desc_parts[0] if desc_parts else ""
                full_description = " | ".join(desc_parts)
                
                # Extract transaction data
                transaction = {
                    "first_name": first_name,
                    "last_name": last_name,
                    "card_number": card_number,
                    "amount": self._clean_amount(amount_value),
                    "transaction_date": self._clean_date(sheet.cell(row=row_num, column=self.column_map.get('transaction_date', 0)).value) if 'transaction_date' in self.column_map else None,
                    "posting_date": self._clean_date(sheet.cell(row=row_num, column=self.column_map.get('business_process_date', 0)).value) if 'business_process_date' in self.column_map else None,
                    "merchant": merchant_name,
                    "description": full_description,
                    "reference_number": self._clean_value(sheet.cell(row=row_num, column=self.column_map.get('transaction_reference', 0)).value) if 'transaction_reference' in self.column_map else "",
                    "row_number": row_num
                }
                
                # Group by cardholder
                if cardholder_name not in transactions_by_cardholder:
                    transactions_by_cardholder[cardholder_name] = []
                
                transactions_by_cardholder[cardholder_name].append(transaction)
            
            wb.close()
            logger.info(f"Parsed {len(transactions_by_cardholder)} cardholders with transactions")
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            raise
        
        return transactions_by_cardholder
    
    def generate_csv_files(self, transactions_by_cardholder: Dict[str, List[Dict]], 
                          output_dir: str, statement_month: int, statement_year: int) -> Dict[str, str]:
        """
        Generate CSV files for each cardholder in Vista import format.
        Returns dict with cardholder names as keys and file paths as values.
        """
        os.makedirs(output_dir, exist_ok=True)
        file_paths = {}
        
        for cardholder_name, transactions in transactions_by_cardholder.items():
            if not transactions:
                continue
            
            # Generate filename
            filename = f"{cardholder_name}.csv"
            filepath = os.path.join(output_dir, filename)
            
            # Calculate total amount
            total_amount = sum(t["amount"] for t in transactions)
            
            # Generate AP reference
            ap_reference = self._generate_ap_reference(
                transactions[0]["first_name"],
                transactions[0]["last_name"],
                statement_month
            )
            
            # Write CSV file
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write APHB header record
                header_row = [
                    "APHB",
                    self.vendor_code,
                    f"{total_amount:.2f}",
                    ap_reference,
                    "",  # Empty columns
                    "",
                    "",
                    "",
                    "",
                    ""
                ]
                writer.writerow(header_row)
                
                # Write APLB line records
                for idx, transaction in enumerate(transactions, 1):
                    line_row = [
                        "APLB",
                        self.record_type,
                        f"{transaction['amount']:.2f}",
                        "",  # GL Account (to be coded)
                        "",  # Empty
                        self.jcco,
                        "",  # Job (to be coded)
                        "",  # Phase (to be coded)
                        "",  # Cost Type (to be coded)
                        f"{transaction['description']} - {transaction['merchant']}"
                    ]
                    writer.writerow(line_row)
            
            file_paths[cardholder_name] = filepath
            logger.info(f"Generated CSV for {cardholder_name}: {filename}")
        
        return file_paths
    
    def generate_coding_template(self, transactions: List[Dict], output_path: str) -> None:
        """Generate a coding template CSV with transaction details."""
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'Transaction Date', 'Posting Date', 'Description', 'Merchant', 
                'Amount', 'GL Account', 'Job Code', 'Phase', 'Cost Type', 'Notes'
            ]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for transaction in transactions:
                writer.writerow({
                    'Transaction Date': transaction['transaction_date'],
                    'Posting Date': transaction['posting_date'],
                    'Description': transaction['description'],
                    'Merchant': transaction['merchant'],
                    'Amount': transaction['amount'],
                    'GL Account': '',  # To be filled by coder
                    'Job Code': '',    # To be filled by coder
                    'Phase': '',       # To be filled by coder
                    'Cost Type': '',   # To be filled by coder
                    'Notes': ''        # To be filled by coder
                })
    
    def _clean_value(self, value):
        """Clean and standardize cell values."""
        if value is None:
            return ""
        return str(value).strip()
    
    def _clean_amount(self, value):
        """Clean and convert amount values."""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        # Handle string amounts with commas
        return float(str(value).replace(',', '').replace('$', ''))
    
    def _clean_date(self, value):
        """Clean and format date values."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        # Try to parse string dates
        try:
            return datetime.strptime(str(value), "%m/%d/%Y")
        except:
            return None
    
    def _generate_ap_reference(self, first_name: str, last_name: str, month: int) -> str:
        """Generate AP reference in format: amex{month}{initials}"""
        month_str = str(month).zfill(2)
        initials = f"{first_name[0]}{last_name[0]}".upper()
        return f"amex{month_str}{initials}"