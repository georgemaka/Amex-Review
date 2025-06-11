import os
import csv
from datetime import datetime
from openpyxl import load_workbook
from typing import Dict, List, Tuple

# Constants
STARTING_ROW = 15
FIRST_NAME_COLUMN = 11
LAST_NAME_COLUMN = 10
AMOUNT_COLUMN = 18
DATE_COLUMN = 16
DESCRIPTION1_COLUMN = 19
DESCRIPTION2_COLUMN = 20

def get_input(prompt: str) -> str:
    return input(prompt)

def create_output_folder(folder_path: str) -> None:
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

def get_filename(row: Tuple) -> str:
    """Generate filename in the format 'firstName lastName.csv'"""
    return f"{row[FIRST_NAME_COLUMN]} {row[LAST_NAME_COLUMN]}.csv"

def process_excel_file(file_path: str) -> Dict[str, float]:
    workbook = load_workbook(file_path)
    sheet = workbook.active
    totals = {}

    for row in sheet.iter_rows(min_row=STARTING_ROW, values_only=True):
        if row[LAST_NAME_COLUMN]:  # Check if the last name column is not empty
            filename = get_filename(row)
            amount = float(str(row[AMOUNT_COLUMN]).replace(',', ''))
            totals[filename] = totals.get(filename, 0) + amount

    return totals

def write_csv_data(output_folder: str, data: Dict[str, List[Dict]], amex_month: str) -> None:
    for filename, items in data.items():
        file_path = os.path.join(output_folder, filename)
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            if items:
                writer = csv.DictWriter(f, fieldnames=items[0].keys())
                writer.writeheader()
                writer.writerows(items)

def process_data(sheet: object, totals: Dict[str, float], amex_month: str) -> Dict[str, List[Dict]]:
    data = {}
    filenames = set()

    for row in sheet.iter_rows(min_row=STARTING_ROW, values_only=True):
        if row[LAST_NAME_COLUMN]:  # Check if the last name column is not empty
            filename = get_filename(row)
            apref_name = f"{row[FIRST_NAME_COLUMN][0]}{row[LAST_NAME_COLUMN][0]}"
            rectype_num = list(totals.keys()).index(filename)

            if filename not in filenames:
                filenames.add(filename)
                data[filename] = []

                # Add APHB record
                aphb_item = {
                    'RecType': 'APHB', 'RecKey': rectype_num, 'Vendor': 19473,
                    'Type': None, 'JCCo': None, 'GLAcct': None, 'Job': None,
                    'Phase': None, 'CostTypeJob': None, 'EquipmentNo': None,
                    'CostCode': None, 'CostType': None, 'InvTotal': round(totals[filename], 2),
                    'Amount': None, 'Reviewer': None, 'BusinessProcessDate': None,
                    'Description': None, 'APRef': f"amex{amex_month}{apref_name}"
                }
                data[filename].append(aphb_item)

            # Add APLB record
            reviewer = f"{row[FIRST_NAME_COLUMN]} {row[LAST_NAME_COLUMN]} AMEX"
            description = f"{row[DESCRIPTION1_COLUMN]} {row[DESCRIPTION2_COLUMN]}"
            aplb_item = {
                'RecType': 'APLB', 'RecKey': rectype_num, 'Vendor': None,
                'Type': 3, 'JCCo': 1, 'GLAcct': None, 'Job': None,
                'Phase': None, 'CostTypeJob': None, 'EquipmentNo': None,
                'CostCode': None, 'CostType': None, 'InvTotal': None,
                'Amount': row[AMOUNT_COLUMN], 'Reviewer': reviewer,
                'BusinessProcessDate': row[DATE_COLUMN], 'Description': description,
                'APRef': None
            }
            data[filename].append(aplb_item)

    return data

def update_filenames(folder_path: str) -> None:
    for filename in os.listdir(folder_path):
        new_filename = filename.title()
        os.rename(os.path.join(folder_path, filename), os.path.join(folder_path, new_filename))

def main():
    input_file = get_input('Input full Statement file path: ')
    output_folder = get_input('Input output folder: ')
    amex_month = get_input('Enter month of Amex data (mmyy): ')

    try:
        create_output_folder(output_folder)

        workbook = load_workbook(input_file)
        sheet = workbook.active
        totals = process_excel_file(input_file)
        data = process_data(sheet, totals, amex_month)

        write_csv_data(output_folder, data, amex_month)

        print('INFO: DONE!')

        os.chdir(output_folder)
        print("Please wait ... Updating Names ")
        update_filenames(output_folder)
        print("All set! Names have been changed.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()