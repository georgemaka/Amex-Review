import os
import win32com.client
from datetime import datetime
from openpyxl import load_workbook
from collections import defaultdict


def get_user_input():
    month = input("Enter the month (e.g., August): ")
    year = input("Enter the year (e.g., 2024): ")

    while True:
        return_date_str = input("Enter the return date (MM/DD/YYYY): ")
        try:
            return_date = datetime.strptime(return_date_str, "%m/%d/%Y")
            formatted_return_date = return_date.strftime("%A, %B %d, %Y")
            break
        except ValueError:
            print("Invalid date format. Please use MM/DD/YYYY.")

    return month, year, formatted_return_date


def create_draft_email(
    outlook, from_email, to_email, cc_email, subject, body, attachments
):
    mail = outlook.CreateItem(0)  # 0 represents olMailItem
    mail.To = to_email
    if cc_email:
        mail.CC = cc_email
    mail.Subject = subject
    mail.HTMLBody = body

    # Set the sender email address
    mail.SentOnBehalfOfName = from_email

    for attachment in attachments:
        if os.path.exists(attachment):
            mail.Attachments.Add(attachment)
    mail.Save()


def main():
    month, year, return_date = get_user_input()

    pdf_folder = r"C:\Users\gmakakaufaki\Sukut\Accounting - Accounts Payable\01 Vendors\A thru H\American Express 19473\2025\05-2025\1009\02 PDF"
    csv_folder = r"C:\Users\gmakakaufaki\Sukut\Accounting - Accounts Payable\01 Vendors\A thru H\American Express 19473\2025\05-2025\1009\01 Raw Imports"

    cardholder_file = r"C:\Users\gmakakaufaki\Sukut\Accounting - Accounts Payable\01 Vendors\A thru H\American Express 19473\CardHolders.xlsx"

    outlook = win32com.client.Dispatch("Outlook.Application")

    missing_files = []

    # Load the Excel workbook
    wb = load_workbook(filename=cardholder_file, read_only=True)
    sheet = wb.active

    # Assuming the first row contains headers
    headers = [cell.value for cell in sheet[1]]

    # Group attachments by email and CC
    email_groups = defaultdict(list)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))

        pdf_name = f"{row_data['PDF']}.pdf"
        csv_name = f"{row_data['csv_name']}.csv"

        pdf_path = os.path.join(pdf_folder, pdf_name)
        csv_path = os.path.join(csv_folder, csv_name)

        # Find the PDF file with the date suffix
        pdf_files = [
            f
            for f in os.listdir(pdf_folder)
            if f.startswith(row_data["PDF"]) and f.endswith(".pdf")
        ]
        if pdf_files:
            pdf_path = os.path.join(pdf_folder, pdf_files[0])
        else:
            missing_files.append(pdf_name)
            continue

        if not os.path.exists(csv_path):
            missing_files.append(csv_name)
            continue

        email_key = (row_data["Code_Email"], row_data["CC_Email"])
        email_groups[email_key].append((pdf_path, csv_path))

    subject = f"{month} {year} American Express Charges"
    body_template = """
    <html>
    <body>
    <p><b>Return By: {return_date}</b></p>
    <p>Attached are the AMEX charges for the month of {month} in Excel and PDF.</p>
    <p>Please code the statement using the CSV format.</p>
    <p>If you have any questions, please reach out to Accounting.</p>
    <p>GL@Sukut.com</p>
    </body>
    </html>
    """

    for (to_email, cc_email), attachments in email_groups.items():
        body = body_template.format(return_date=return_date, month=month)
        try:
            create_draft_email(
                outlook,
                "GL@sukut.com",  # From email
                to_email,
                cc_email,
                subject,
                body,
                [
                    path for pair in attachments for path in pair
                ],  # Flatten the list of attachment pairs
            )
            print(f"Draft email created for {to_email}")
        except Exception as e:
            print(f"Error creating draft email for {to_email}: {str(e)}")

    wb.close()

    if missing_files:
        print("The following files were missing:")
        for file in missing_files:
            print(file)
    else:
        print("All draft emails created successfully.")


if __name__ == "__main__":
    main()
