import os
import win32com.client
from datetime import datetime
from openpyxl import load_workbook
from collections import defaultdict

def get_user_input():
    month = input("Enter the month (e.g., August): ")
    year = input("Enter the year (e.g., 2024): ")
    return month, year

def create_draft_email(outlook, from_email, to_emails, subject, body, attachments):
    mail = outlook.CreateItem(0)  # 0 represents olMailItem
    mail.To = "; ".join(to_emails)
    mail.Subject = subject
    mail.HTMLBody = body
    mail.SentOnBehalfOfName = from_email

    for attachment in attachments:
        if os.path.exists(attachment):
            mail.Attachments.Add(attachment)
    mail.Save()

def main():
    month, year = get_user_input()

    pdf_folder = r"C:\Users\gmakakaufaki\Sukut\Accounting - Accounts Payable\01 Vendors\A thru H\American Express 19473\2025\03-2025\02 PDF"
    reviewer_file = r"C:\Users\gmakakaufaki\Sukut\Accounting - Accounts Payable\01 Vendors\A thru H\American Express 19473\CardHolders-Reviewers.xlsx"

    outlook = win32com.client.Dispatch("Outlook.Application")

    missing_files = []

    # Load the Excel workbook
    wb = load_workbook(filename=reviewer_file, read_only=True)
    sheet = wb.active

    # Assuming the first row contains headers
    headers = [cell.value for cell in sheet[1]]
    
    # Group PDFs by reviewers
    reviewer_groups = defaultdict(list)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))

        pdf_name = f"{row_data['PDF']}"
        
        # Find the PDF file with the date suffix
        pdf_files = [f for f in os.listdir(pdf_folder) if f.startswith(pdf_name) and f.endswith('.pdf')]
        if pdf_files:
            pdf_path = os.path.join(pdf_folder, pdf_files[0])
        else:
            missing_files.append(f"{pdf_name}.pdf")
            continue

        reviewers = [row_data.get(f'Reviewer{i}', '') for i in range(1, 4)]
        reviewers = [r for r in reviewers if r]  # Remove empty entries

        for reviewer in reviewers:
            reviewer_groups[reviewer].append(pdf_path)

    subject = f"{month} {year} American Express Statement Review"
    body_template = """
    <html>
    <body>
    <p>Attached are the American Express statements for the month of {month} for your review.</p>
    <p>If you have any questions, please reach out to Accounting.</p>
    <p>GL@Sukut.com</p>
    </body>
    </html>
    """

    for reviewer_email, attachments in reviewer_groups.items():
        body = body_template.format(month=month)
        try:
            create_draft_email(
                outlook,
                "GL@sukut.com",  # From email
                [reviewer_email],
                subject,
                body,
                attachments
            )
            print(f"Draft email created for {reviewer_email}")
        except Exception as e:
            print(f"Error creating draft email for {reviewer_email}: {str(e)}")

    wb.close()

    if missing_files:
        print("The following files were missing:")
        for file in missing_files:
            print(file)
    else:
        print("All draft emails created successfully.")

if __name__ == "__main__":
    main()